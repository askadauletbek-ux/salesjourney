from __future__ import annotations

import datetime
import base64
import hashlib
import hmac
import io
import json
import os
import time
from typing import Any, Dict, Optional, Tuple
from collections import defaultdict
from functools import wraps
import urllib.parse

import requests
from flask import (
    Blueprint,
    current_app,
    jsonify,
    redirect,
    render_template,
    request,
    session,
    send_file,
    abort, url_for
)
from flask_login import current_user, login_required  # <--- Добавлен login_required
from sqlalchemy import select, and_  # <--- Добавлены select и and_

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
except ImportError:
    Workbook = None

from extensions import db
# Обновлен список импортируемых моделей
from models import Company, AmoCRMConnection, AmoCRMUserMap, PartnerUser, User, Challenge, ChallengeProgress, ChallengeGoalType, ChallengeMode, UserRole, AmoCRMUserDailyStat
# --- Blueprints ---
bp_amocrm_company_api = Blueprint("amocrm_company_api", __name__, url_prefix="/api/partners/company")
bp_amocrm_pages = Blueprint("amocrm_pages", __name__, url_prefix="/partner/company")

# --- Constants ---
WON_STATUS_ID = 142
LOST_STATUS_ID = 143
AMO_STATE_SECRET = os.getenv("AMO_STATE_SECRET", "dev_secret_change_me_please")
AMO_REDIRECT_BASE = os.getenv("AMO_REDIRECT_BASE", "http://localhost:5000")


# --- Decorators & Auth ---

def partner_owns_company_required(f):
    @wraps(f)
    def wrapper(company_id, *args, **kwargs):
        # 1. Проверяем, залогинен ли пользователь вообще
        if not current_user.is_authenticated:
            return current_app.login_manager.unauthorized()

        # 2. Проверяем, есть ли у него профиль партнера
        p = current_user.partner_profile
        if not p:
            # Пользователь залогинен, но он не партнер (например, обычный сотрудник)
            return abort(403)

        c = db.session.get(Company, company_id)
        if not c:
            return abort(404)

        # 3. Проверяем, владеет ли партнер этой компанией
        if c.owner_partner_id != p.id:
            return abort(403)

        return f(company_id=company_id, *args, **kwargs)

    return wrapper


def _refresh_if_needed(company_id: int) -> Optional[AmoCRMConnection]:
    conn = _get_connection_or_none(company_id)
    if not conn or not conn.refresh_token:
        return conn

    now = int(time.time())
    if conn.expires_at and now + 60 < conn.expires_at:
        return conn

    try:
        token_url = f"https://{conn.base_domain}/oauth2/access_token"
        payload = {
            "client_id": conn.client_id,
            "client_secret": conn.client_secret,
            "grant_type": "refresh_token",
            "refresh_token": conn.refresh_token,
            "redirect_uri": _callback_url(company_id),
        }
        r = requests.post(token_url, json=payload, timeout=15)
        if r.status_code == 200:
            _save_tokens(conn, r.json())
            return conn
        else:
            current_app.logger.warning("AMO refresh failed for company %d: %s %s", company_id, r.status_code, r.text)
            if r.status_code in [400, 401]:
                _clear_tokens(company_id)
            return None
    except requests.RequestException as e:
        current_app.logger.exception("AMO refresh request exception for company %d", company_id)
        abort(503, description=f"Не удалось подключиться к AmoCRM для обновления токена: {e}")


def _amo_get(base_domain: str, access_token: str, path: str, params: Dict[str, Any]) -> requests.Response:
    url = f"https://{base_domain}{path}"
    headers = {"Authorization": f"Bearer {access_token}", "Content-Type": "application/json"}
    try:
        return requests.get(url, headers=headers, params=params, timeout=60)
    except requests.RequestException as e:
        current_app.logger.exception("AmoCRM API GET request failed")
        abort(503, description=f"Не удалось выполнить запрос к API AmoCRM: {e}")


# --- Utils ---
def _b64url(data: bytes) -> str:
    return base64.urlsafe_b64encode(data).decode().rstrip("=")


def _sign_state(payload: dict) -> str:
    raw = json.dumps(payload, separators=(",", ":"), ensure_ascii=False).encode()
    sig = hmac.new(AMO_STATE_SECRET.encode(), raw, hashlib.sha256).digest()
    return _b64url(sig)


def _get_connection_or_none(company_id: int) -> Optional[AmoCRMConnection]:
    return AmoCRMConnection.query.filter_by(company_id=company_id).first()


def _save_tokens(conn: AmoCRMConnection, data: Dict[str, Any]) -> None:
    conn.access_token = data.get("access_token")
    conn.refresh_token = data.get("refresh_token")
    conn.expires_at = int(time.time()) + int(data.get("expires_in", 0))
    if "base_domain" in data:
        conn.base_domain = data["base_domain"]
    conn.last_sync_at = int(time.time())
    db.session.add(conn)
    db.session.commit()


def _clear_tokens(company_id: int) -> None:
    conn = _get_connection_or_none(company_id)
    if conn:
        db.session.delete(conn)
        db.session.commit()


def _callback_url(company_id: int = None) -> str:
    """
    Генерирует глобальный Callback URL.
    Аргумент company_id оставлен для совместимости, но не используется в URL.
    """
    if os.getenv("AMO_REDIRECT_URI"):
        return os.getenv("AMO_REDIRECT_URI")

    # Генерируем ссылку на новый глобальный роут
    path = url_for('amocrm_company_api.global_amocrm_callback')
    return f"{AMO_REDIRECT_BASE.rstrip('/')}{path}"

def _partner_company_url(company_id: int) -> str:
    return url_for("amocrm_pages.company_crm_page", company_id=company_id)


def _fetch_users_map(base_domain: str, access_token: str) -> Dict[int, Dict[str, Any]]:
    out: Dict[int, Dict[str, Any]] = {}
    page, limit = 1, 250
    while True:
        r = _amo_get(base_domain, access_token, "/api/v4/users", {"page": page, "limit": limit})
        if r.status_code != 200:
            current_app.logger.warning("AMO users fetch %s: %s", r.status_code, r.text)
            break
        data = r.json() or {}
        users = (data.get("_embedded") or {}).get("users") or []
        for u in users:
            uid = u.get("id")
            if uid is not None:
                out[int(uid)] = {
                    "id": int(uid),
                    "name": u.get("name") or f"User {uid}",
                    "email": (u.get("email") or "") if isinstance(u.get("email"), str) else "",
                }
        if len(users) < limit:
            break
        page += 1
    return out


def _iter_closed_leads(base_domain: str, access_token: str, ts_from: int, ts_to: int):
    page, limit = 1, 250
    while True:
        params = {
            "page": page,
            "limit": limit,
            "filter[closed_at][from]": ts_from,
            "filter[closed_at][to]": ts_to,
            "order[closed_at]": "desc",
        }
        r = _amo_get(base_domain, access_token, "/api/v4/leads", params)
        if r.status_code != 200:
            current_app.logger.error("AMO leads fetch %s: %s", r.status_code, r.text)
            break
        data = r.json() or {}
        leads = (data.get("_embedded") or {}).get("leads") or []
        if not leads:
            break
        for lead in leads:
            yield lead
        if len(leads) < limit:
            break
        page += 1


def _iter_created_leads(base_domain: str, access_token: str, ts_from: int, ts_to: int):
    page, limit = 1, 250
    while True:
        params = {
            "page": page,
            "limit": limit,
            "filter[created_at][from]": ts_from,
            "filter[created_at][to]": ts_to,
            "order[created_at]": "desc",
        }
        r = _amo_get(base_domain, access_token, "/api/v4/leads", params)
        if r.status_code != 200:
            current_app.logger.error("AMO leads(created) fetch %s: %s", r.status_code, r.text)
            break
        data = r.json() or {}
        leads = (data.get("_embedded") or {}).get("leads") or []
        if not leads:
            break
        for lead in leads:
            yield lead
        if len(leads) < limit:
            break
        page += 1


def _period_from_request() -> Tuple[int, int, int, str]:
    now = datetime.datetime.now()
    rng = (request.args.get("range") or "this_week").lower().strip()
    a_from = request.args.get("from")
    a_to = request.args.get("to")

    # Пользовательский диапазон дат
    if rng == "custom" and a_from and a_to:
        try:
            ts_from = int(a_from)
            ts_to = int(a_to)
            dt_to = datetime.datetime.fromtimestamp(ts_to).replace(hour=23, minute=59, second=59)
            ts_to = int(dt_to.timestamp())
            days = max(0, int((ts_to - ts_from) / 86400))
            return ts_from, ts_to, days, "custom"
        except (ValueError, TypeError):
            pass

    if rng == "today":
        start_of_day = now.replace(hour=0, minute=0, second=0, microsecond=0)
        return int(start_of_day.timestamp()), int(now.timestamp()), 0, "today"

    if rng == "this_week":
        start_of_week = now - datetime.timedelta(days=now.weekday())
        start_of_week = start_of_week.replace(hour=0, minute=0, second=0, microsecond=0)
        return int(start_of_week.timestamp()), int(now.timestamp()), 7, "this_week"

    if rng == "last_week":
        end_of_last_week = now - datetime.timedelta(days=now.weekday() + 1)
        end_of_last_week = end_of_last_week.replace(hour=23, minute=59, second=59, microsecond=999999)
        start_of_last_week = end_of_last_week - datetime.timedelta(days=6)
        start_of_last_week = start_of_last_week.replace(hour=0, minute=0, second=0, microsecond=0)
        return int(start_of_last_week.timestamp()), int(end_of_last_week.timestamp()), 7, "last_week"

    if rng == "prev_last_week":
        end_of_prev_last_week = now - datetime.timedelta(days=now.weekday() + 8)
        end_of_prev_last_week = end_of_prev_last_week.replace(hour=23, minute=59, second=59, microsecond=999999)
        start_of_prev_last_week = end_of_prev_last_week - datetime.timedelta(days=6)
        start_of_prev_last_week = start_of_prev_last_week.replace(hour=0, minute=0, second=0, microsecond=0)
        return int(start_of_prev_last_week.timestamp()), int(end_of_prev_last_week.timestamp()), 7, "prev_last_week"

    start_of_week = now - datetime.timedelta(days=now.weekday())
    start_of_week = start_of_week.replace(hour=0, minute=0, second=0, microsecond=0)
    return int(start_of_week.timestamp()), int(now.timestamp()), 7, "this_week"


def _compute_stats(base_domain: str, access_token: str, ts_from: int, ts_to: int) -> Dict[str, Any]:
    users_map = _fetch_users_map(base_domain, access_token)
    by_user = defaultdict(lambda: {"created": 0, "won": 0, "lost": 0})

    for lead in _iter_created_leads(base_domain, access_token, ts_from, ts_to):
        resp_id = lead.get("responsible_user_id") or 0
        by_user[resp_id]["created"] += 1

    for lead in _iter_closed_leads(base_domain, access_token, ts_from, ts_to):
        resp_id = lead.get("responsible_user_id") or 0
        status_id = int(lead.get("status_id") or 0)
        if status_id == WON_STATUS_ID:
            by_user[resp_id]["won"] += 1
        elif status_id == LOST_STATUS_ID:
            by_user[resp_id]["lost"] += 1

    rows = []
    total_created = 0
    total_won = 0
    total_lost = 0
    all_user_ids = by_user.keys()

    for uid in all_user_ids:
        stats = by_user[uid]
        created, won, lost = stats["created"], stats["won"], stats["lost"]
        conversion = round(100 * won / created) if created > 0 else 0
        name = (users_map.get(uid) or {}).get("name") if uid else "Без ответственного"
        rows.append({
            "user_id": uid,
            "display_name": name,
            "created": created,
            "won": won,
            "lost": lost,
            "conv": conversion
        })
        total_created += created
        total_won += won
        total_lost += lost

    overall_conversion = round(100 * total_won / total_created) if total_created > 0 else 0

    return {
        "created_count": total_created,
        "won_count": total_won,
        "lost_count": total_lost,
        "overall_conversion": overall_conversion,
        "rows": rows,
    }


def _apply_view_filters(rows: list[Dict[str, Any]], sort: str, min_total: int, q: str) -> list[Dict[str, Any]]:
    if q:
        ql = q.lower()
        rows = [r for r in rows if (r.get("display_name") or "").lower().find(ql) >= 0]
    if min_total > 0:
        rows = [r for r in rows if (r.get("won", 0) + r.get("lost", 0)) >= min_total]
    if sort == "conv_desc":
        rows.sort(key=lambda r: (r.get("conv", 0), r.get("won", 0)), reverse=True)
    elif sort == "lost_asc":
        rows.sort(key=lambda r: (r.get("lost", 0), -r.get("won", 0)))
    elif sort == "name_asc":
        rows.sort(key=lambda r: str(r.get("display_name") or ""))
    else:  # won_desc
        rows.sort(key=lambda r: (-r.get("won", 0), r.get("lost", 0)))
    return rows


# ======== API Endpoints ========

# --- Challenge API ---

@bp_amocrm_company_api.route("/<int:company_id>/challenges", methods=["GET"])
@partner_owns_company_required
def list_challenges(company_id: int):
    """Список всех челленджей компании (для админки)"""
    challenges = db.session.execute(
        select(Challenge).where(Challenge.company_id == company_id).order_by(Challenge.end_date.desc())
    ).scalars().all()

    items = []
    for c in challenges:
        items.append({
            "id": c.id,
            "name": c.name,
            "start": c.start_date.isoformat(),
            "end": c.end_date.isoformat(),
            "type": c.goal_type.value,
            "goal": c.goal_value,
            "mode": c.mode.value,
            "active": c.is_active
        })
    return jsonify({"items": items})


@bp_amocrm_company_api.route("/<int:company_id>/challenges", methods=["POST"])
@partner_owns_company_required
def create_challenge(company_id: int):
    """Создание нового челленджа"""
    data = request.get_json(silent=True) or {}

    try:
        new_challenge = Challenge(
            company_id=company_id,
            name=data["name"],
            description=data.get("description", ""),
            start_date=datetime.date.fromisoformat(data["start_date"]),
            end_date=datetime.date.fromisoformat(data["end_date"]),
            goal_type=ChallengeGoalType(data["goal_type"]),
            goal_value=int(data["goal_value"]),
            mode=ChallengeMode(data.get("mode", "PERSONAL"))
        )
        db.session.add(new_challenge)
        db.session.commit()
        return jsonify({"ok": True, "id": new_challenge.id})
    except Exception as e:
        return jsonify({"error": str(e)}), 400


@bp_amocrm_company_api.route("/<int:company_id>/challenges/active_leaderboard")
@login_required
def active_challenge_leaderboard(company_id: int):
    """Возвращает данные для Лидерборда и Общей Цели (Squad Goal)"""
    # Доступ для сотрудников компании или владельца
    if current_user.company_id != company_id and current_user.role not in [UserRole.PARTNER, UserRole.COMPANY_OWNER]:
        abort(403)

    today = datetime.date.today()

    # 1. Находим активный челлендж (берем ближайший к завершению)
    challenge = db.session.execute(
        select(Challenge).where(
            and_(
                Challenge.company_id == company_id,
                Challenge.is_active == True,
                Challenge.start_date <= today,
                Challenge.end_date >= today
            )
        ).order_by(Challenge.end_date.asc())
    ).scalars().first()

    if not challenge:
        return jsonify({"active": False})

    # 2. Собираем прогресс
    results = db.session.execute(
        select(ChallengeProgress, User)
        .join(User, ChallengeProgress.user_id == User.id)
        .where(ChallengeProgress.challenge_id == challenge.id)
        .order_by(ChallengeProgress.current_value.desc())
    ).all()

    leaderboard = []
    total_progress = 0

    for progress, user in results:
        total_progress += progress.current_value
        leaderboard.append({
            "user_id": user.id,
            "name": user.username,
            "value": progress.current_value,
        })

    # Расчет Squad Goal
    team_percent = 0
    if challenge.goal_value > 0:
        if challenge.mode == ChallengeMode.TEAM:
            # Для командного режима: сумма всех / цель
            team_percent = int((total_progress / challenge.goal_value) * 100)
        else:
            # Для личного: можно показать прогресс лидера
            pass

    return jsonify({
        "active": True,
        "challenge": {
            "name": challenge.name,
            "type": challenge.goal_type.value,
            "mode": challenge.mode.value,
            "target": challenge.goal_value,
            "team_progress_val": total_progress,
            "team_progress_percent": min(100, team_percent)
        },
        "leaderboard": leaderboard
    })


@bp_amocrm_company_api.route("/<int:company_id>/crm/amocrm/status")
@partner_owns_company_required
def amocrm_status(company_id: int):
    conn = _refresh_if_needed(company_id)
    if not conn or not conn.access_token:
        return jsonify({"connected": False})
    return jsonify({
        "connected": True,
        "base_domain": conn.base_domain,
        "token_expires_at": conn.expires_at,
        "last_sync_at": conn.last_sync_at,
    })


@bp_amocrm_company_api.route("/<int:company_id>/crm/amocrm/unlink", methods=["POST"])
@partner_owns_company_required
def amocrm_unlink(company_id: int):
    _clear_tokens(company_id)
    return jsonify({"ok": True})


@bp_amocrm_company_api.route("/<int:company_id>/crm/amocrm/connect", methods=["POST"])
@partner_owns_company_required
def amocrm_connect(company_id: int):
    data = request.get_json(silent=True) or request.form
    client_id = (data.get("client_id") or "").strip()
    client_secret = (data.get("client_secret") or "").strip()
    base_domain = (data.get("base_domain") or "").strip()

    if not all([client_id, client_secret, base_domain]):
        return jsonify({"error": "client_id, client_secret и base_domain обязательны"}), 400

    conn = _get_connection_or_none(company_id)
    if not conn:
        conn = AmoCRMConnection(company_id=company_id)
    conn.client_id = client_id
    conn.client_secret = client_secret
    conn.base_domain = base_domain
    db.session.add(conn)
    db.session.commit()

    payload = {"cid": company_id, "ts": int(time.time())}
    state_obj = {"p": payload, "s": _sign_state(payload)}
    state = _b64url(json.dumps(state_obj, separators=(",", ":"), ensure_ascii=False).encode())

    # Используем новую функцию для получения Redirect URI
    redirect_uri = _callback_url()

    params = {
        "client_id": client_id,
        "response_type": "code",
        "redirect_uri": redirect_uri,
        "state": state,
        "mode": "post_message",
    }

    auth_url = f"https://www.amocrm.ru/oauth?{urllib.parse.urlencode(params, safe=':/')}"
    return jsonify({"auth_url": auth_url})


@bp_amocrm_company_api.route("/common/callback")
def global_amocrm_callback():
    """
    Единая точка входа для всех OAuth-ответов.
    Не содержит ID компании в URL, извлекает его из параметра state.
    """
    code = request.args.get("code")
    state_from_req = request.args.get("state")
    referer_domain = request.args.get("referer", "").strip()

    if not code or not state_from_req:
        return "Invalid OAuth callback: missing code or state", 400

    # 1. Распаковка state и поиск company_id
    try:
        raw = base64.urlsafe_b64decode(state_from_req + "==")
        parsed = json.loads(raw.decode())
        payload = parsed.get("p", {})
        sig = parsed.get("s", "")

        # Проверка подписи
        if _sign_state(payload) != sig:
            raise ValueError("Invalid state signature")

        # Извлекаем ID компании из state
        company_id = int(payload.get("cid"))

    except Exception as e:
        current_app.logger.warning("Invalid AmoCRM callback state: %s", e)
        return "Bad state or signature validation failed", 400

    # 2. Получение соединения
    conn = _get_connection_or_none(company_id)
    if not conn:
        return f"Connection not configured for company {company_id}", 400

    if referer_domain:
        conn.base_domain = referer_domain
        db.session.add(conn)
        db.session.commit()

    if not conn.base_domain:
        return "Connection domain not configured", 400

    # 3. Обмен кода на токен
    redirect_uri = _callback_url()  # Используем тот же общий URL
    token_url = f"https://{conn.base_domain}/oauth2/access_token"

    data = {
        "client_id": conn.client_id,
        "client_secret": conn.client_secret,
        "grant_type": "authorization_code",
        "code": code,
        "redirect_uri": redirect_uri,
    }

    try:
        r = requests.post(token_url, json=data, timeout=15)
        if r.status_code != 200:
            current_app.logger.error("AmoCRM token exchange failed: %s %s", r.status_code, r.text)
            return f"Token exchange failed with status {r.status_code}. Response: {r.text}", 400
        _save_tokens(conn, r.json())
    except requests.RequestException as e:
        current_app.logger.exception("AmoCRM token exchange error for company %d: %s", company_id, e)
        return "Token request error", 500

    # 4. Редирект обратно на страницу CRM конкретной компании
    return render_template("oauth_callback_close.html", redirect_url=_partner_company_url(company_id))


# Старый роут оставляем для истории или удаляем, если уверены, что никто не использует старые ссылки
@bp_amocrm_company_api.route("/<int:company_id>/crm/amocrm/callback")
def amocrm_callback(company_id: int):
    return redirect(url_for('amocrm_company_api.global_amocrm_callback', **request.args))
    code = request.args.get("code")
    state_from_req = request.args.get("state")
    referer_domain = request.args.get("referer", "").strip()

    if not code or not state_from_req:
        return "Invalid OAuth callback: missing code or state", 400

    try:
        raw = base64.urlsafe_b64decode(state_from_req + "==")
        parsed = json.loads(raw.decode())
        payload = parsed.get("p", {})
        sig = parsed.get("s", "")
        if _sign_state(payload) != sig or int(payload.get("cid")) != company_id:
            raise ValueError("Invalid state signature or company ID")
    except Exception as e:
        current_app.logger.warning("Invalid AmoCRM callback state: %s", e)
        return "Bad state", 400

    conn = _get_connection_or_none(company_id)
    if not conn:
        return "Connection not configured", 400

    if referer_domain:
        conn.base_domain = referer_domain
        db.session.add(conn)
        db.session.commit()

    if not conn.base_domain:
        return "Connection domain not configured", 400

    redirect_uri = os.getenv("AMO_REDIRECT_URI") or (
                AMO_REDIRECT_BASE.rstrip("/") + url_for("amocrm_company_api.amocrm_callback", company_id=company_id))
    token_url = f"https://{conn.base_domain}/oauth2/access_token"

    data = {
        "client_id": conn.client_id,
        "client_secret": conn.client_secret,
        "grant_type": "authorization_code",
        "code": code,
        "redirect_uri": redirect_uri,
    }

    try:
        r = requests.post(token_url, json=data, timeout=15)
        if r.status_code != 200:
            current_app.logger.error("AmoCRM token exchange failed: %s %s", r.status_code, r.text)
            return f"Token exchange failed with status {r.status_code}. Response: {r.text}", 400
        _save_tokens(conn, r.json())
    except requests.RequestException as e:
        current_app.logger.exception("AmoCRM token exchange error for company %d: %s", company_id, e)
        return "Token request error", 500

    return render_template("oauth_callback_close.html", redirect_url=_partner_company_url(company_id))


@bp_amocrm_company_api.route("/my/sync_stats", methods=["POST"])
@login_required
def sync_my_daily_stats():
    """
    SIPUNI FIX V2: Проверка ответственного внутри примечания + Debug Log.
    """
    user = current_user
    today = datetime.date.today()

    if not user.company_id:
        return jsonify({"error": "No company"}), 400

    mapping = db.session.execute(
        select(AmoCRMUserMap).where(
            and_(
                AmoCRMUserMap.platform_user_id == user.id,
                AmoCRMUserMap.company_id == user.company_id
            )
        )
    ).scalar_one_or_none()

    if not mapping:
        return jsonify({"linked": False, "message": "Account not linked"}), 200

    conn = _refresh_if_needed(user.company_id)
    if not conn or not conn.access_token:
        return jsonify({"error": "Company integration not active"}), 400

    debug_log = []  # Лог для отладки

    try:
        # 1. Время (UTC - 3 часа)
        start_of_day_utc = int(datetime.datetime.combine(today, datetime.time.min).timestamp())
        start_of_day_safe = start_of_day_utc - 10800

        my_amo_id = mapping.amocrm_user_id
        debug_log.append(f"My AmoID: {my_amo_id}, Time: {start_of_day_safe}")

        # --- 2. СБОР ЗВОНКОВ ---
        calls_count = 0
        talk_seconds = 0

        # Запрашиваем ВСЕ события (без фильтра по типу, чтобы не упустить)
        params_events = {
            "filter[created_at][from]": start_of_day_safe,
            "limit": 100,
            "with": "note"  # Просим расширить данные заметкой, если это заметка
        }

        r_ev = _amo_get(conn.base_domain, conn.access_token, "/api/v4/events", params_events)

        if r_ev.status_code == 200:
            events = r_ev.json().get("_embedded", {}).get("events", [])
            debug_log.append(f"Total events found: {len(events)}")

            for ev in events:
                # Пытаемся достать данные примечания
                vals_after = ev.get("value_after", [])

                # Нормализация данных (иногда список, иногда словарь)
                wrapper = vals_after[0] if isinstance(vals_after, list) and vals_after else vals_after
                if isinstance(wrapper, list): wrapper = {}  # Защита от странных форматов

                # Ищем структуру note
                note_data = wrapper.get("note", {})

                # Если в событии нет note, пропускаем
                if not note_data:
                    continue

                # --- ПРОВЕРКА ПРИНАДЛЕЖНОСТИ ---
                # 1. Ответственный за саму заметку (Это главное для SIPUNI)
                note_resp_id = int(note_data.get("responsible_user_id") or 0)
                # 2. Создатель события (резервный вариант)
                event_creator_id = int(ev.get("created_by") or 0)

                # Если ни ответственный, ни создатель не мы - пропускаем
                if my_amo_id != note_resp_id and my_amo_id != event_creator_id:
                    continue

                # --- ПРОВЕРКА ТИПА ---
                note_type = str(note_data.get("note_type", ""))

                # Для отладки сохраним первый найденный note для нашего юзера
                if len(debug_log) < 5:
                    debug_log.append(f"Check Note: Type={note_type}, Resp={note_resp_id}")

                if note_type in ["call_in", "call_out", "10", "11", "12", "13"]:
                    # Достаем длительность
                    params = note_data.get("params", {})
                    dur_val = params.get("duration")

                    if dur_val is not None:
                        calls_count += 1
                        try:
                            talk_seconds += int(dur_val)
                        except:
                            pass
        else:
            debug_log.append(f"API Error: {r_ev.status_code}")

        # --- 3. СБОР СДЕЛОК ---
        leads_created = 0
        leads_won = 0
        leads_lost = 0

        # Created
        r_cr = _amo_get(conn.base_domain, conn.access_token, "/api/v4/leads", {
            "filter[created_at][from]": start_of_day_safe,
            "filter[responsible_user_id]": my_amo_id,
            "limit": 250
        })
        if r_cr.status_code == 200:
            leads_created = len(r_cr.json().get("_embedded", {}).get("leads", []))

        # Closed
        r_cl = _amo_get(conn.base_domain, conn.access_token, "/api/v4/leads", {
            "filter[closed_at][from]": start_of_day_safe,
            "filter[responsible_user_id]": my_amo_id,
            "limit": 250
        })
        if r_cl.status_code == 200:
            closed = r_cl.json().get("_embedded", {}).get("leads", [])
            for l in closed:
                sid = int(l.get("status_id", 0))
                if sid == WON_STATUS_ID:
                    leads_won += 1
                elif sid == LOST_STATUS_ID:
                    leads_lost += 1

        # --- 4. СОХРАНЕНИЕ ---
        stat_entry = db.session.execute(
            select(AmoCRMUserDailyStat).where(
                and_(AmoCRMUserDailyStat.user_id == user.id, AmoCRMUserDailyStat.date == today)
            )
        ).scalar_one_or_none()

        if not stat_entry:
            stat_entry = AmoCRMUserDailyStat(user_id=user.id, date=today)
            db.session.add(stat_entry)

        stat_entry.calls_count = calls_count
        stat_entry.talk_seconds = talk_seconds
        stat_entry.leads_created = leads_created
        stat_entry.leads_won = leads_won
        stat_entry.leads_lost = leads_lost

        db.session.commit()

        # Возвращаем DEBUG LOG в JSON, чтобы вы могли его прислать мне
        return jsonify({
            "linked": True,
            "calls": stat_entry.calls_count,
            "minutes": stat_entry.minutes_talked,
            "conversion": stat_entry.conversion,
            "updated_at_str": "Только что",
            "debug_log": debug_log  # <--- Посмотрите это поле в Network
        })

    except Exception as e:
        return jsonify({"error": "Sync failed", "details": str(e), "debug_log": debug_log}), 500

@bp_amocrm_company_api.route("/<int:company_id>/crm/amocrm/sync", methods=["POST"])
@partner_owns_company_required
def amocrm_sync(company_id: int):
    conn = _get_connection_or_none(company_id)
    if not conn:
        return jsonify({"error": "not connected"}), 400
    conn.last_sync_at = int(time.time())
    db.session.add(conn)
    db.session.commit()
    return jsonify({"ok": True})


@bp_amocrm_company_api.route("/<int:company_id>/crm/stats")
@partner_owns_company_required
def crm_stats(company_id: int):
    conn = _refresh_if_needed(company_id)
    if not conn or not conn.access_token:
        return jsonify({"error": "not connected or token invalid"}), 400

    ts_from, ts_to, days, label = _period_from_request()
    data = _compute_stats(conn.base_domain, conn.access_token, ts_from, ts_to)

    return jsonify({
        "range": label,
        "from": ts_from,
        "to": ts_to,
        "days": days,
        "created_count": data["created_count"],
        "won_count": data["won_count"],
        "lost_count": data["lost_count"],
        "conversion": data["overall_conversion"],
        "by_user": data["rows"],
    })


@bp_amocrm_company_api.route("/<int:company_id>/crm/stats.xlsx")
@partner_owns_company_required
def crm_stats_xlsx(company_id: int):
    conn = _refresh_if_needed(company_id)
    if not conn or not conn.access_token:
        return jsonify({"error": "not connected or token invalid"}), 400

    ts_from, ts_to, days, label = _period_from_request()
    base = _compute_stats(conn.base_domain, conn.access_token, ts_from, ts_to)

    sort = request.args.get("sort", "won_desc")
    min_total = int(request.args.get("min_total", "0"))
    q = request.args.get("q", "").strip()

    rows = _apply_view_filters(list(base["rows"]), sort, min_total, q)
    filename_label = "today" if label == "today" else (f"{days}d" if days else "custom")
    fname = f"crm_stats_company_{company_id}_{filename_label}.xlsx"

    if Workbook is None:
        output = io.StringIO()
        output.write("\ufeff")
        output.write("ID;Пользователь;Успешно;Не реализовано;Конверсия (%)\n")
        for r in rows:
            output.write(f'{r["user_id"]};{r["display_name"]};{r["won"]};{r["lost"]};{r["conv"]}\n')
        mem = io.BytesIO(output.getvalue().encode("utf-8"))
        mem.seek(0)
        return send_file(
            mem,
            mimetype="text/csv; charset=utf-8",
            as_attachment=True,
            download_name=f"crm_stats_company_{company_id}_{filename_label}.csv",
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "Статистика"
    ws["A1"] = "Статистика по сотрудникам"
    ws["A2"] = f"Период: {'сегодня' if label == 'today' else f'последние {days} дней' if days else 'задан вручную'}"
    ws["A3"] = f"Сформировано: {time.strftime('%Y-%m-%d %H:%M:%S', time.localtime())}"
    ws.merge_cells("A1:E1")
    ws.merge_cells("A2:E2")
    ws.merge_cells("A3:E3")

    # Стили
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")

    headers = ["ID", "Пользователь", "Успешно", "Не реализовано", "Конверсия (%)"]
    for col_num, h in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col_num, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for idx, r in enumerate(rows, 6):
        ws.cell(row=idx, column=1, value=r["user_id"])
        ws.cell(row=idx, column=2, value=r["display_name"])
        ws.cell(row=idx, column=3, value=r["won"])
        ws.cell(row=idx, column=4, value=r["lost"])
        ws.cell(row=idx, column=5, value=r["conv"])

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return send_file(
        bio,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=fname,
    )


@bp_amocrm_company_api.route("/<int:company_id>/crm/users")
@partner_owns_company_required
def crm_users(company_id: int):
    conn = _refresh_if_needed(company_id)
    if not conn or not conn.access_token:
        return jsonify({"connected": False, "error": "not connected"}), 400

    users = list(_fetch_users_map(conn.base_domain, conn.access_token).values())
    return jsonify({"connected": True, "users": users})


@bp_amocrm_company_api.route("/<int:company_id>/crm/map/list")
@partner_owns_company_required
def crm_map_list(company_id: int):
    rows = AmoCRMUserMap.query.filter_by(company_id=company_id).all()
    mapping = {str(r.platform_user_id): r.amocrm_user_id for r in rows}
    return jsonify({"map": mapping, "count": len(mapping)})


@bp_amocrm_company_api.route("/<int:company_id>/crm/map", methods=["POST"])
@partner_owns_company_required
def crm_map_set(company_id: int):
    body = request.get_json(silent=True) or {}
    platform_id = body.get("platform_user_id")
    amocrm_id = body.get("amocrm_user_id")
    if platform_id is None or amocrm_id is None:
        return jsonify({"error": "platform_user_id and amocrm_user_id required"}), 400

    row = AmoCRMUserMap.query.filter_by(company_id=company_id, platform_user_id=platform_id).first()
    if row:
        row.amocrm_user_id = amocrm_id
    else:
        row = AmoCRMUserMap(company_id=company_id, platform_user_id=platform_id, amocrm_user_id=amocrm_id)
    db.session.add(row)
    db.session.commit()
    return jsonify({"ok": True})


@bp_amocrm_company_api.route("/<int:company_id>/crm/rt")
@partner_owns_company_required
def crm_realtime(company_id: int):
    conn = _refresh_if_needed(company_id)
    if not conn or not conn.access_token:
        return jsonify({"error": "not connected"}), 400

    now = int(time.time())
    lt = time.localtime(now)
    midnight = int(time.mktime((lt.tm_year, lt.tm_mon, lt.tm_mday, 0, 0, 0, lt.tm_wday, lt.tm_yday, lt.tm_isdst)))

    won_by = defaultdict(int)
    lost_by = defaultdict(int)
    for lead in _iter_closed_leads(conn.base_domain, conn.access_token, midnight, now):
        uid = lead.get("responsible_user_id") or 0
        sid = int(lead.get("status_id") or 0)
        if sid == WON_STATUS_ID:
            won_by[uid] += 1
        elif sid == LOST_STATUS_ID:
            lost_by[uid] += 1

    created_by = defaultdict(int)
    for lead in _iter_created_leads(conn.base_domain, conn.access_token, midnight, now):
        uid = lead.get("responsible_user_id") or 0
        created_by[uid] += 1

    users_map = _fetch_users_map(conn.base_domain, conn.access_token)

    total_won = sum(won_by.values())
    total_lost = sum(lost_by.values())
    total_created = sum(created_by.values())
    total_all = total_won + total_lost
    conversion = round(100 * total_won / total_all) if total_all else 0
    hours_passed = max(1.0, (now - midnight) / 3600.0)

    per_user = []
    user_ids = set(list(won_by.keys()) + list(lost_by.keys()) + list(created_by.keys()))
    for uid in user_ids:
        w, l, c = int(won_by.get(uid, 0)), int(lost_by.get(uid, 0)), int(created_by.get(uid, 0))
        tot = w + l
        conv = round(100 * w / tot) if tot else 0
        vph = round(w / hours_passed, 2)
        meta = users_map.get(int(uid), {})
        per_user.append({
            "amocrm_user_id": int(uid),
            "amocrm_name": meta.get("name") or f"User {uid}",
            "amocrm_email": meta.get("email") or "",
            "won": w, "lost": l, "created": c, "conv": conv, "wins_per_hour": vph,
        })

    per_user.sort(key=lambda r: (-r["won"], -r["conv"], -r["created"]))

    return jsonify({
        "range": "today", "from": midnight, "to": now, "updated_at": now,
        "kpi": {
            "won_today": total_won, "lost_today": total_lost,
            "created_today": total_created, "conversion_today": conversion,
        },
        "users": per_user,
    })


# ===== Pages =====
@bp_amocrm_pages.route("/<int:company_id>/crm")
@partner_owns_company_required
def company_crm_page(company_id: int):
    c = db.session.get(Company, company_id)
    slug = c.slug if c else ""
    return render_template("partner_company_crm.html", company_id=company_id, company_slug=slug)


@bp_amocrm_pages.route("/<int:company_id>/crm/dashboard")
@partner_owns_company_required
def company_crm_dashboard(company_id: int):
    c = db.session.get(Company, company_id)
    slug = c.slug if c else ""
    # Передаем invite_code в шаблон
    return render_template("partner_company_crm_dashboard.html", company_id=company_id, company_slug=slug,
                           invite_code=c.invite_code)


@bp_amocrm_company_api.route("/<int:company_id>/debug/inspect_entity")
@partner_owns_company_required
def debug_inspect_entity(company_id: int):
    """
    ПОЛНЫЙ РЕНТГЕН СУЩНОСТИ:
    Выгружает Calls, Notes и Events, привязанные к ID сделки или контакта.
    Позволяет увидеть, в каком именно поле SIPUNI прячет данные о звонках.

    Использование:
    /api/partners/company/{id}/debug/inspect_entity?entity_id={ID_СДЕЛКИ}&type=leads
    или
    /api/partners/company/{id}/debug/inspect_entity?entity_id={ID_КОНТАКТА}&type=contacts
    """
    entity_id = request.args.get('entity_id')
    # type может быть: leads (сделки), contacts (контакты)
    entity_type = request.args.get('type', 'leads')

    if not entity_id:
        return jsonify({"error": "Укажите ?entity_id=..."}), 400

    conn = _refresh_if_needed(company_id)
    if not conn or not conn.access_token:
        return jsonify({"error": "Not connected"}), 400

    report = {
        "target": f"{entity_type} #{entity_id}",
        "raw_calls_api": [],  # Стандартные звонки
        "raw_notes_api": [],  # Примечания (ОЧЕНЬ ВАЖНО для SIPUNI)
        "raw_events_api": []  # События таймлайна
    }

    try:
        # 1. API CALLS (Звонки, привязанные к сущности)
        params_calls = {
            "filter[entity_id]": entity_id,
            "filter[entity_type]": entity_type,
            "with": "duration"
        }
        r_calls = _amo_get(conn.base_domain, conn.access_token, "/api/v4/calls", params_calls)
        if r_calls.status_code == 200:
            report["raw_calls_api"] = r_calls.json().get("_embedded", {}).get("calls", [])

        # 2. API NOTES (Примечания - сюда часто пишут интеграции)
        # Получаем список примечаний конкретной сущности
        url_notes = f"/api/v4/{entity_type}/{entity_id}/notes"
        params_notes = {"limit": 50, "order[created_at]": "desc"}
        r_notes = _amo_get(conn.base_domain, conn.access_token, url_notes, params_notes)

        if r_notes.status_code == 200:
            notes = r_notes.json().get("_embedded", {}).get("notes", [])
            # Сохраняем как есть, чтобы вы увидели структуру
            report["raw_notes_api"] = notes

        # 3. API EVENTS (События таймлайна)
        params_events = {
            "filter[entity_id]": entity_id,
            "filter[entity_type]": entity_type,
            "limit": 50,
            "order[created_at]": "desc"
        }
        r_events = _amo_get(conn.base_domain, conn.access_token, "/api/v4/events", params_events)
        if r_events.status_code == 200:
            report["raw_events_api"] = r_events.json().get("_embedded", {}).get("events", [])

        return jsonify(report)

    except Exception as e:
        return jsonify({"error": str(e)}), 500

@bp_amocrm_company_api.route("/<int:company_id>/members")
@partner_owns_company_required
def company_members(company_id: int):
    """
    Возвращает список сотрудников, зарегистрированных в компании (через код).
    """
    c = db.session.get(Company, company_id)
    if not c:
        return jsonify({"items": []})

    # Берем всех юзеров с ролью EMPLOYEE/MANAGER, привязанных к этой компании
    members = []
    for u in c.employees:
        members.append({
            "id": u.id,
            "display_name": u.username,
            "email": u.email,
            "role": u.role.value
        })
    return jsonify({"items": members})